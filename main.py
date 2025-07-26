from openpyxl import load_workbook
import time
# 定义常量
INPUT_DIR = "/mnt/sharedisk/班级学习日详情.xlsx"
COMPLETE = 100
PASS = 60
unpassName = []
incompleteName = []
unpassPer = []
incompletePer = []
'''
contentStyle = openpyxl.styles.Font(bold=True, size=10)
titleStyleA = openpyxl.styles.Font(name="Arial", size=11, bold=True)
titleStyleB = openpyxl.styles.Alignment(horizontal="center", vertical="center")  # 水平居中
titleBackground = openpyxl.styles.PatternFill(fill_type="solid", start_color="0080AAFF", end_color="0080AAFF")  # 标题背景色
'''
# 准备工作表
inputWorkbook = load_workbook(INPUT_DIR)
sheet = inputWorkbook.active
sheet.title = "MAIN SHEET"
# 获取数据,名称和完成率

cellRange_Name = sheet["C3":"C55"]
cellRange_completePercentage = sheet["H3":"H55"]


def getData(cellRange):
    Data = []
    for location in cellRange:
        cellLocation = str(location)[20:][0:-3]
        Data.append(sheet[cellLocation].value)
    return Data


name = getData(cellRange_Name)
completePercentage = getData(cellRange_completePercentage)
# 接下来检查完成率，并将不合格的标记出来
step = 0
for per in completePercentage:
    per = float(per)
    if per < COMPLETE and per >= PASS:
        incompleteName.append(name[step])
        incompletePer.append(str(per))

    if per < PASS:
        unpassName.append(name[step])
        unpassPer.append(str(per))
    step += 1
print("未完成名单：", incompleteName,"\n",incompletePer)
print("完成率小于60：", unpassName, "\n", unpassPer)

formatted_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
# 以markdown形式输出结果ft
with open("/mnt/sharedisk/result.md", "w") as file:
    index_ = 0
    file.write("# " + formatted_time + "\n")
    file.write("-------" + "\n" + "\n")
    file.write("|姓名|完成率|" + "\n")
    file.write("|:-----:|:-----:|" + "\n")
    for a in incompleteName:
        if incompleteName == []:
            break
        file.write("|" + incompleteName[index_])
        file.write("|" + incompletePer[index_] + "|" + "\n")
        index_ += 1
    index_ = 0
    for a in unpassName:
        if unpassName == []:
            break
        file.write("|" + unpassName[index_])
        file.write("|" + unpassPer[index_] + "|" + "\n")
        index_ += 1
